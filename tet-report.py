from tetpyclient import RestClient
import tetpyclient
import json
import requests.packages.urllib3
import sys
import os
import argparse
import time
import csv
from columnar import columnar
from time import mktime
from datetime import datetime
import datetime
from argparse import ArgumentParser
from collections import defaultdict
from tqdm import tqdm as progress
import urllib3
import xlsxwriter
from xlsxwriter import Workbook
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart3D,Reference 
import re



CEND = "\33[0m"     #End
CGREEN = "\33[32m"  #Information
CYELLOW = "\33[33m" #Request Input
CRED = "\33[31m"    #Error
URED = "\33[4;31m" 
Cyan = "\33[0;36m"  #Return
BLINK = "\33[5m"
BOLD = "\33[1m"
ITALIC = "\33[3m"
UNDERLINE = "\33[4m"
LBLUE = "\33[1;34m"

# =================================================================================
# feedback: Le Anh Duc - anhdle@cisco.com
# See reason below -- why verify=False param is used
# python3 tet-report.py --url https://10.71.129.30/ --credential Japan_api_credentials.json
# =================================================================================
requests.packages.urllib3.disable_warnings()


parser = argparse.ArgumentParser(description='Tetration Get all sensors')
parser.add_argument('--url', help='Tetration URL', required=True)
parser.add_argument('--credential', help='Path to Tetration json credential file', required=True)
args = parser.parse_args()

# =================================================================================
# Overall
# =================================================================================
def CreateRestClient():
    rc = RestClient(args.url,
                    credentials_file=args.credential, verify=False)
    return rc


# =================================================================================
# Report
# =================================================================================
def GetApplicationScopes(rc):
    resp = rc.get('/app_scopes')

    if resp.status_code != 200:
        print("Failed to retrieve app scopes")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowScopes(scopes):
    """
        List all the Scopes in Tetration Appliance
        Scope ID | Scope Name | Parent Scope | VRF | Policy Priority
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Number', 'Scope ID', 'Name', 'Parent Scope', 'VRF', 'Policy Priority']
        data_list = [[i+1, x['id'],
                    x['name'],
                    x['parent_app_scope_id'],
                    x['vrf_id'], x['policy_priority']] for i,x in enumerate(scopes) ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetSensors(rc):
    resp = rc.get('/sensors')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve sensors list")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowAgents(sensors):
    """
        List all the agents registered in Tetration Appliance
        Hostname | Agent Type | Last checkin | Install Date | Version | Scopes
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Number', 'Host Name', 'UUID', 'Agent Type', 'Last Check-in', 'Install Date', 'Version', 'Scopes']
        data_list = [[i+1, x['host_name'], x['uuid'], x['agent_type'], time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(x['last_config_fetch_at'])), time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x['created_at'])), x['current_sw_version'], ','.join(set([y['vrf'] for y in x['interfaces']])) ]for i,x in enumerate(sensors['results']) ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def ShowAgentProfile(agent):
    """
        Detail of an agent
        """
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Host Name', 'Agent Type', 'Last Check-in', 'Platform', 'Version', 'Scopes']
        data_list = [[agent['host_name'], agent['agent_type'], time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(agent['last_config_fetch_at'])), agent['platform'], agent['current_sw_version'], ','.join(set([y['vrf'] for y in agent['interfaces']]))]]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetFlowMetrics(rc):
    resp = rc.get('/flowsearch/metrics')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve metrics list"+ CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def selectAgent(sensors):
    # Return UUID for one Sensors that we choose
    print (Cyan + "\nHere are all Software Sensors in your cluster: " + CEND)
    ShowAgents(sensors)
    choice = input('\nSelect which Sensor (Number) above you want to know detail: ')
    return sensors['results'][int(choice)-1]['uuid']


def GetAgentProfile(rc,uuid):
    resp = rc.get('/workload/' + uuid)

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetWorkloadStats(rc,uuid, t0, t1, td):
    #td = 15 * 60 # 15 minutes
    resp = rc.get('/workload/' + uuid + '/stats?t0=' + str(t0) + '&t1=' + str(t1) + '&td=' + str(td))

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowWorkloadStats(stats):
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Time', 'Flow Count', 'received bytes', 'received packets', 'transmitted bytes', 'transmitted packets']
        data_list = [[x['timestamp'], x['result']['flow_count'], x['result']['rx_byte_count'],  x['result']['rx_packet_count'], x['result']['tx_byte_count'], x['result']['tx_packet_count']]for x in stats ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetSwPackages(rc,uuid):
    resp = rc.get('/workload/' + uuid + '/packages')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetApps(rc):
    resp = rc.get('/applications')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Apps list" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowApps(Apps):
    AppsList = []
    headers = ['Number', 'App Name', 'Author', 'App ID', 'Primary?']
    for i,app in enumerate(Apps): AppsList.append([i+1,app["name"] , app['author'], app["id"], app['primary']])
    table = columnar(AppsList, headers, no_borders=False)
    print(table)

def GetAppVersions(rc, appid):
    resp = rc.get('/applications/' + appid + '/versions')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve Apps list" + CEND)
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetLatestVersion(app_versions):
    version =[]
    try:
        for vers in app_versions: 
            if 'v' in vers["version"]: version.append(vers['version'])
        return version[0]
    except:
        print(URED + "Failed to retrieve latest app version"+ CEND)

def downloadConvs(rc,appIDs):
    # Download Apps Conversation JSON files from Apps workspace
    apps = []
    limit = int(input ("How many conversation you want to download? "))
    for appID in appIDs:
        print('Downloading app details for '+appID + "into json file")
        versions = GetAppVersions(rc,appID)
        version = int(re.search(r'\d+', GetLatestVersion(versions)).group(0))
        req_payload = {"version": version,
               "limit": limit
               }
        resp = rc.post('/openapi/v1/conversations/%s'%appID, json_body=json.dumps(req_payload))
        if resp.status_code == 200:
            parsed_resp = json.loads(resp.content)
            apps.append(parsed_resp)
    
    with open('all-conversations.json', "w") as config_file:
                json.dump(apps, config_file, indent=4)
                print("all-conversations.json created")



def ShowConversationTet(convs):
    """
        Show All conversation and export to Excel file
        Source IP | Source Filter Name | Destination IP | Destination Filter Name | Protocol | Port | Bytes | Packets
        """
    data_list = []
    headers = ['Source IP', 'Destination IP', 'Protocol', 'Port', 'Bytes', 'Packets']
    listconv = convs[0]
    for x in listconv['results']:
        data_list.append([x['src_ip'], x['dst_ip'], x['protocol'], x['port'], x['byte_count'], x['packet_count']]) 
    table = columnar(data_list, headers, no_borders=False)
    print(table)
    with open('conversation.csv', 'w') as csvfile:
        writer = csv.writer(csvfile, delimiter=',')
        writer.writerow(i for i in headers)
        for row in data_list:
            writer.writerow(row)
    
    export_xlsfile = 'Apps_Conversation.xlsx'
    workbook = xlsxwriter.Workbook(export_xlsfile)
    bold = workbook.add_format({'bold': True})
    worksheet = workbook.add_worksheet(name='Apps Conversation')
    header_format = workbook.add_format()
    header_format.set_bg_color('cyan')
    header_format.set_bold()
    header_format.set_font_size(13)
    header_format.set_font_color('black')
    worksheet.set_row(0, None)
    worksheet.write_row(0,0,headers,header_format)
    i=1
    firstline = True
    with open('conversation.csv', 'r') as f:
        for row in csv.reader(f):
            if firstline:    #skip first line
                firstline = False
                continue
            worksheet.write_row(i,0,row)
            i += 1
    worksheet.set_column(0, 0, 20)
    worksheet.set_column(1, 1, 20)
    i =2
    while i < 6:
        worksheet.set_column(i, i, 15)
        i += 1
    workbook.close()
    print ('Writing csv file to %s with %d columns' % (export_xlsfile, len(headers)))
    os.remove('conversation.csv')


def ShowSwPackages(packages):
    columns = None
    if columns:
            headers = []
            data_list = []
    else:
        headers = ['Name', 'Architecture', 'Publisher', 'Version']
        data_list = [[x['name'], x['architecture'], x['publisher'],  x['version']]for x in packages ]
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetVul(rc,uuid):
    resp = rc.get('/workload/' + uuid + '/vulnerabilities')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowVul(vuls):
    data_list = []
    headers = ['Package Information', 'CVE ID', 'v2 Score', 'v3 Score', 'v2_severity', 'v2_access_complexity', 'v3_base_severity', 'v3_attack_complexity']
    search_key = 'v3_score'
    for x in vuls: 
        if search_key in x.keys(): data_list.append([x['package_infos'], x['cve_id'], x['v2_score'], x['v3_score'], x['v2_severity'], x['v2_access_complexity'],  x['v3_base_severity'], x['v3_attack_complexity']])
        else: data_list.append([x['package_infos'], x['cve_id'], x['v2_score'], 'None', x['v2_severity'], x['v2_access_complexity'],  'None' , 'None'])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetProc(rc,uuid):
    resp = rc.get('/workload/' + uuid + '/process/list')

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowProc(proc):
    data_list = []
    headers = ['PID', 'PPID', 'Proc State', 'Username','CMD', 'Exec_Path', 'Package Name', 'Package Version']
    search_key = 'pkg_info_name'
    for x in proc['ps_row']:
        if search_key in x.keys(): data_list. append([x['pid'], x['ppid'], x['proc_state'],  x['username'], x['cmd'], x['exec_path'], x['pkg_info_name'],  x['pkg_info_version']])
        else: data_list. append([x['pid'], x['ppid'], x['proc_state'], x['username'], x['cmd'], x['exec_path'], 'NA',  'NA'])
    table = columnar(data_list, headers, no_borders=False)
    print(table)

def GetProcTree(rc,uuid):
    payload = {}
    resp = rc.post('/workload/' + uuid + '/process/tree/ids', json_body=json.dumps(payload))

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def GetProcTreeDetail(rc,uuid, handle):
    payload = {"handle": handle}
    resp = rc.post('/workload/' + uuid + '/process/tree/details', json_body=json.dumps(payload))

    if resp.status_code != 200:
        print(URED + "Failed to retrieve agent detail")
        print(resp.status_code)
        print(resp.text)
    else:
        return resp.json()

def ShowProcTreeDetail(procDetail):
    data_list = []
    headers = ['PID', 'PPID', 'Proc State', 'Username','CMD', 'Exec_Path', 'Package Name', 'Package Version']
    search_key = 'pkg_info_name'
    for x in procDetail['results']:
        if search_key in x.keys(): data_list. append([x['process_id'], x['parent_process_id'], x['proc_state'],  x['username'], x['command_string'], x['exec_path'], x['pkg_info_name'],  x['pkg_info_version']])
        else: data_list. append([x['process_id'], x['parent_process_id'], x['proc_state'], x['username'], x['command_string'], x['exec_path'], 'NA',  'NA'])
    table = columnar(data_list, headers, no_borders=False)
    print(table)


def get_inventory(rc, end_point, req_payload):
    '''
    Get the list of inventory items matching the query
    '''

    all_result = []

    resp = rc.post(end_point, json_body=json.dumps(req_payload))
    results = resp.json()

    all_result += results["results"]

    while results.get("offset"):
        # Get the offset ID for page 2
        next_page = results["offset"]
        # Set the offset to page 2
        req_payload["offset"] = next_page

        resp = rc.post(end_point, json_body=json.dumps(req_payload))
        results = resp.json()

        all_result += results["results"]

    return all_result


def get_inventory_cve(rc):

    criticality = int(input("Which CVE Score you want to query your inventory (from 0 to 10): "))
    req_payload = {"filter": {"type": "or",
                              "filters": [{"type": "gt", "field": "host_tags_cvss2", "value": criticality},
                                          {"type": "gt", "field": "host_tags_cvss3", "value": criticality}]}}

#    req_payload = {'filter': {"type": "eq", "field": "ip", "value": "192.168.2.98"}}


    cve_hosts = get_inventory(rc, '/inventory/search', req_payload)

    #print (json.dumps(cve_hosts, indent=4))

    cve_list = []   # store host data with CVE info
    print (BLINK + CRED + 'Processing vulnerabilities data ........ ' + CEND)

    for host in cve_hosts:

        host_uuid = str(host["host_uuid"])

        host_name = host['host_name']

        results = GetVul(rc,host_uuid)

        #print (CYELLOW + 'Gathering CVE data for ' + host_name + " with UUID " + host_uuid + CEND)

        #print (json.dumps(results, indent=4))

        for pkg in results:
            cve_dict = {}
            if "v2_score" in pkg.keys():
                if "v3_score" in pkg.keys():
                    if (int(pkg["v2_score"]) > criticality) or int(pkg["v3_score"]) > criticality :
                        cve_dict["IP"] = host["ip"]
                        cve_dict["Hostname"] = host["host_name"]
                        cve_dict["OS"] = host["os"]
                        cve_dict["Version"] = host["os_version"]
                        cve_dict["Package Info"] = pkg["package_infos"]
                        cve_dict["Scope"] = host["tags_scope_name"]
                        cve_dict["CVE ID"] = pkg["cve_id"]
                   
                        cve_dict["CVE v2 Score"] = pkg["v2_score"]
                        cve_dict["CVE v2 Severity"] = pkg["v2_severity"]
                        cve_dict["CVE v2 access vector"] = pkg["v2_access_vector"]
                        cve_dict["CVE v2 access complexity"] = pkg["v2_access_complexity"]

                        cve_dict["CVE v3 Score"] = pkg["v3_score"]
                        cve_dict["CVE v3 Severity"] = pkg["v3_base_severity"]
                        cve_dict["CVE v3 attack vector"] = pkg["v3_attack_vector"]
                        cve_dict["CVE v3 attack complexity"] = pkg["v3_attack_complexity"]
                        cve_dict["CVE v3 availability impact"] = pkg["v3_availability_impact"]
                else:
                    if (int(pkg["v2_score"]) > criticality):
                        cve_dict["IP"] = host["ip"]
                        cve_dict["Hostname"] = host["host_name"]
                        cve_dict["OS"] = host["os"]
                        cve_dict["Version"] = host["os_version"]
                        cve_dict["Package Info"] = pkg["package_infos"]
                        cve_dict["Scope"] = host["tags_scope_name"]
                        cve_dict["CVE ID"] = pkg["cve_id"]
                   
                        cve_dict["CVE v2 Score"] = pkg["v2_score"]
                        cve_dict["CVE v2 Severity"] = pkg["v2_severity"]
                        cve_dict["CVE v2 access vector"] = pkg["v2_access_vector"]
                        cve_dict["CVE v2 access complexity"] = pkg["v2_access_complexity"]
            else:
                if (int(pkg["v3_score"]) > criticality):
                    cve_dict["IP"] = host["ip"]
                    cve_dict["Hostname"] = host["host_name"]
                    cve_dict["OS"] = host["os"]
                    cve_dict["Version"] = host["os_version"]
                    cve_dict["Package Info"] = pkg["package_infos"]
                    cve_dict["Scope"] = host["tags_scope_name"]
                    cve_dict["CVE ID"] = pkg["cve_id"]
               
                    cve_dict["CVE v3 Score"] = pkg["v3_score"]
                    cve_dict["CVE v3 Severity"] = pkg["v3_base_severity"]
                    cve_dict["CVE v3 attack vector"] = pkg["v3_attack_vector"]
                    cve_dict["CVE v3 attack complexity"] = pkg["v3_attack_complexity"]
                    cve_dict["CVE v3 availability impact"] = pkg["v3_availability_impact"]

            cve_list.append(cve_dict)

            cve_list_final = []
            for string in cve_list:
                if (string != ""): cve_list_final.append(string)

    # specify csv file for exporting
    export_xlsfile = 'cve_hosts_final.xlsx'
    export_csvfile = 'cve_hosts_final.csv'
    temp_csv = 'cve_hosts.csv'

    # specify csv header fields
    csv_header = ["IP", "Hostname", "OS", "Version", "Package Info", "Scope", "CVE ID", 
                  "CVE v2 Score", "CVE v2 Severity", "CVE v2 access vector", "CVE v2 access complexity",
                  "CVE v3 Score", "CVE v3 Severity", "CVE v3 attack vector", "CVE v3 attack complexity", "CVE v3 availability impact"]

    
    # Export file in csv format
    with open(temp_csv, 'w+') as f:
        writer = csv.DictWriter(f, csv_header, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        for row in cve_list_final:
            writer.writerow(row)

    with open('cve_hosts.csv') as infile, open('cve_hosts_final.csv', 'w', newline='') as output:
         writer = csv.writer(output)
         for row in csv.reader(infile):
             if any(field.strip() for field in row):
                 writer.writerow(row)

    workbook = xlsxwriter.Workbook(export_xlsfile)
    bold = workbook.add_format({'bold': True})
    worksheet = workbook.add_worksheet(name='CVE Report')
    header_format = workbook.add_format()
    header_format.set_bg_color('cyan')
    header_format.set_bold()
    header_format.set_text_wrap()
    header_format.set_font_size(13)
    header_format.set_font_color('black')
    cell_format = workbook.add_format()
    cell_format.set_text_wrap()
    worksheet.set_row(0, None)
    worksheet.write_row(0,0,csv_header,header_format)
    i=1
    firstline = True
    with open('cve_hosts_final.csv', 'r') as f:
        for row in csv.reader(f):
            if firstline:    #skip first line
                firstline = False
                continue
            worksheet.write_row(i,0,row)
            i += 1
    worksheet.set_column(0, 0, 15)
    worksheet.set_column(1, 1, 15)
    worksheet.set_column(2, 2, 15,cell_format)
    worksheet.set_column(4, 4, 30,cell_format)
    worksheet.set_column(5, 5, 30,cell_format)
    i =6
    while i < 16:
        worksheet.set_column(i, i, 15)
        i += 1
    workbook.close()
    print ('Writing csv file to %s with %d columns' % (export_xlsfile, len(csv_header)))
    os.remove(temp_csv)
    os.remove(export_csvfile)


def get_inventory_flow(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to get inventory statistic: ')
    scope_name = scopes[int(choice)-1]['name']
    subnet = input (CYELLOW + "Which subnet (X.X.X.X/Y) of inventory you want to query: " +CEND)
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),int(from_day),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),int(to_day),0,0).timestamp())
    # Query inventories in the scope
    req_payload = {
    "filter":
            {
                "type": "subnet",
                "field": "ip",
                "value": subnet
            },
    "scopeName": scope_name }

#    req_payload = {'filter': {"type": "eq", "field": "ip", "value": "192.168.2.98"}}

    hosts_in_scope = get_inventory(rc, '/inventory/search', req_payload)

    talkative_list = []   # store host data with bytes
    print (BLINK + CRED +'Processing flow data ........ '+ CEND)

    for host in hosts_in_scope:
        req_endpoint = '/inventory/' + str(host["ip"]) + '-' + str(host["vrf_id"] + '/stats?t0=' + str(t0) +'&t1='+str(t1)+'&td=day')
        #print (req_endpoint)
        
        results = rc.get(req_endpoint).json()

        #print ('Getting conversation data from ' + req_endpoint)
        for x in results:
            stats_dict = {}
            stats_dict["Hostname"] = host["host_name"]
            stats_dict["IP"] = host["ip"]
            stats_dict["Timestamp"] = x["timestamp"]
            stats_dict["OS"] = host["os"]
            stats_dict["OS Version"] = host["os_version"]
            stats_dict["MAC Address"] = host["iface_mac"]
            stats_dict["Received Bytes"] = x["result"]["rx_byte_count"]
            stats_dict["Transmited Bytes"] = x["result"]["tx_byte_count"]
            stats_dict["Total Flows"] = x["result"]["flow_count"]
            stats_dict["Received Packets"] = x["result"]["rx_packet_count"]
            stats_dict["Transmited Packets"] = x["result"]["tx_packet_count"]
        talkative_list.append(stats_dict)

    # specify csv file for exporting
    export_csvfile = 'stats_hosts.xlsx'

    # specify csv header fields
    csv_header = ["Hostname", "IP", "Timestamp", "OS", "OS Version", "MAC Address", "Received Bytes", "Transmited Bytes",
                  "Total Flows", "Received Packets", "Transmited Packets"]

    workbook = xlsxwriter.Workbook(export_csvfile)
    bold = workbook.add_format({'bold': True})
    worksheet = workbook.add_worksheet(name='Subnet Top talkers')
    cell_format = workbook.add_format()
    cell_format.set_bg_color('cyan')
    cell_format.set_bold()
    cell_format.set_font_color('black')
    worksheet.set_row(0, None)
    worksheet.write_row(0,0,csv_header,cell_format)
    i=1
    for row in talkative_list:
        #print (row.values())
        worksheet.write_row(i,0,row.values())
        i+=1
    worksheet.set_column(0, 0, 18)
    worksheet.set_column(1, 1, 15)
    worksheet.set_column(2, 2, 22)
    i =3
    while i < 12:
        worksheet.set_column(i, i, 15)
        i += 1
    workbook.close()

    print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def get_flow_topTalkers(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to query Top Talkers: ')
    scope_name = scopes[int(choice)-1]['name']
    threshold = input('\nHow many top talkers you want to query (Max is 1000): ')
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
    metrics = GetFlowMetrics(rc)
    print (Cyan + "Here are the available metrics: \n" + json.dumps(metrics, indent=4, sort_keys=True) + CEND)
    metric = input (Cyan + "which one you want to query? (copy and paste here): " +  CEND)
    req_payload = {
    "t0": t0,    
    "t1": t1,    
    "dimension": "src_address",
    "metric": metric,
    #"filter": {"type": "eq", "field": "src_address", "value": "172.29.203.193"},  #optional
    "threshold": int(threshold),
    "scopeName": scope_name
    }


    resp = rc.post('/flowsearch/topn',
               json_body=json.dumps(req_payload))

    if resp.status_code != 200:
        print(URED + "Failed to retrieve TopN")
        print(resp.status_code)
        print(resp.text)
    else:
        topN = resp.json()
        print (json.dumps(topN, indent=4))
        topN_list = []   # store TopN data
        print (json.dumps(topN[0]['result'], indent=4))
        for top in topN[0]['result']:
            topN_dict = {}
            topN_dict["Source Address"] = top["src_address"]
            topN_dict[metric] = top[metric]
            topN_list.append(topN_dict)

        # specify csv file for exporting
        export_csvfile = 'topTalkerReport.xlsx'

        # specify csv header fields
        csv_header = ["Source Address", metric]

        
        
        workbook = xlsxwriter.Workbook(export_csvfile)
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet(name='Top Source Address')
        #chartsheet = workbook.add_chartsheet(name='Top Source Address Chart') 
        cell_format = workbook.add_format()
        cell_format.set_bg_color('cyan')
        cell_format.set_bold()
        cell_format.set_font_color('black')
        worksheet.set_row(0, None)
        worksheet.write_row(0,0,csv_header,cell_format)
        i=1
        for row in topN_list:
            #print (row.values())
            worksheet.write_row(i,0,row.values())
            i+=1
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        workbook.close()

        '''workbook = xlsxwriter.Workbook(export_csvfile)
        worksheet = workbook.active 
        chart1 = BarChart3D()
        chart1.title = 'Top Source Address'
        chart1.y_axis.title = 'Source Address'
        chart1.x_axis.title = metric

        data = Reference(worksheet, min_col = 2, min_row = 2, 
                         max_col = 2, max_row = len(topN_list))
        titles = Reference(worksheet, min_col=1, min_row=2, max_row = len(topN_list))
        chart1.add_data(data, titles_from_data=True)
        chart.set_categories(titles)
        chart1.shape = 4
        worksheet.add_chart(chart1, "D07")
        workbook.save(export_csvfile)
        workbook.close()'''

        print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def get_flow_topDest(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to query Top Destination: ')
    scope_name = scopes[int(choice)-1]['name']
    threshold = input('\nHow many top Destination you want to query (Max is 1000): ')
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
    #dimensions = GetFlowDimensions(rc)
    metrics = GetFlowMetrics(rc)
    print (Cyan + "Here are the available metrics: \n" + json.dumps(metrics, indent=4, sort_keys=True) + CEND)
    metric = input (Cyan + "which one you want to query? (copy and paste here): " +  CEND)
    #print (Cyan + "Here are the available dimensions: \n" + json.dumps(metrics, indent=4, sort_keys=True))
    #dimension = input (Cyan + "which one you want to query? (copy and paste here) ")
    req_payload = {
    "t0": t0,    
    "t1": t1,    
    "dimension": "dst_address",
    "metric": metric,
    #"filter": {"type": "eq", "field": "src_address", "value": "172.29.203.193"},  #optional
    "threshold": int(threshold),
    "scopeName": scope_name
    }


    resp = rc.post('/flowsearch/topn',
               json_body=json.dumps(req_payload))

    #print (json.dumps(cve_hosts, indent=4))
    if resp.status_code != 200:
        print(URED + "Failed to retrieve TopN")
        print(resp.status_code)
        print(resp.text)
    else:
        topN = resp.json()
        print (json.dumps(topN, indent=4))
        topN_list = []   # store TopN data
        print (json.dumps(topN[0]['result'], indent=4))
        for top in topN[0]['result']:
            topN_dict = {}
            topN_dict["Destination Address"] = top["dst_address"]
            topN_dict[metric] = top[metric]
            topN_list.append(topN_dict)

        # specify csv file for exporting
        export_csvfile = 'topDestinationReport.xlsx'

        # specify csv header fields
        csv_header = ["Destination Address", metric]

        
        
        workbook = xlsxwriter.Workbook(export_csvfile)
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet(name='Top Destination Address')
        cell_format = workbook.add_format()
        cell_format.set_bg_color('cyan')
        cell_format.set_bold()
        cell_format.set_font_color('black')
        worksheet.set_row(0, None)
        worksheet.write_row(0,0,csv_header,cell_format)
        i=1
        for row in topN_list:
            #print (row.values())
            worksheet.write_row(i,0,row.values())
            i+=1
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        workbook.close()

        print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def get_flow_topDestService(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to query Top Server Service: ')
    scope_name = scopes[int(choice)-1]['name']
    threshold = input('\nHow many Top Server Service you want to query (Max is 1000): ')
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
    #dimensions = GetFlowDimensions(rc)
    metrics = GetFlowMetrics(rc)
    print (Cyan + "Here are the available metrics: \n" + json.dumps(metrics, indent=4, sort_keys=True) + CEND)
    metric = input (Cyan + "which one you want to query? (copy and paste here): " +  CEND)
    #print (Cyan + "Here are the available dimensions: \n" + json.dumps(metrics, indent=4, sort_keys=True))
    #dimension = input (Cyan + "which one you want to query? (copy and paste here) ")
    req_payload = {
    "t0": t0,    
    "t1": t1,    
    "dimension": "dst_port",
    "metric": metric,
    #"filter": {"type": "eq", "field": "src_address", "value": "172.29.203.193"},  #optional
    "threshold": int(threshold),
    "scopeName": scope_name
    }


    resp = rc.post('/flowsearch/topn',
               json_body=json.dumps(req_payload))

    #print (json.dumps(cve_hosts, indent=4))
    if resp.status_code != 200:
        print(URED + "Failed to retrieve TopN")
        print(resp.status_code)
        print(resp.text)
    else:
        topN = resp.json()
        print (json.dumps(topN, indent=4))
        topN_list = []   # store TopN data
        print (json.dumps(topN[0]['result'], indent=4))
        for top in topN[0]['result']:
            topN_dict = {}
            topN_dict["Destination Service"] = top["dst_port"]
            topN_dict[metric] = top[metric]
            topN_list.append(topN_dict)

        # specify csv file for exporting
        export_csvfile = 'topDestinationPort.xlsx'

        # specify csv header fields
        csv_header = ["Destination Service", metric]

        
        
        workbook = xlsxwriter.Workbook(export_csvfile)
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet(name='Top Destination Service')
        cell_format = workbook.add_format()
        cell_format.set_bg_color('cyan')
        cell_format.set_bold()
        cell_format.set_font_color('black')
        worksheet.set_row(0, None)
        worksheet.write_row(0,0,csv_header,cell_format)
        i=1
        for row in topN_list:
            #print (row.values())
            worksheet.write_row(i,0,row.values())
            i+=1
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        workbook.close()
        print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def get_flow_topSrcService(rc):
    rc = CreateRestClient()
    scopes = GetApplicationScopes(rc)
    print (CGREEN + "Here is all scopes in your cluster: " + CEND)
    ShowScopes(scopes)
    choice = input('\nSelect which Scope (Number) bove you want to query Top Client Service: ')
    scope_name = scopes[int(choice)-1]['name']
    threshold = input('\nHow many Top Client Service you want to query (Max is 1000): ')
    from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
    from_month = input(CYELLOW + "Month (mm)? "+CEND)
    from_day = input(CYELLOW + "Day (dd)? "+CEND)
    to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
    to_month = input(CYELLOW + "Month (mm)? "+CEND)
    to_day = input(CYELLOW + "Day (dd)? "+CEND)
    t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
    t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
    #dimensions = GetFlowDimensions(rc)
    metrics = GetFlowMetrics(rc)
    print (Cyan + "Here are the available metrics: \n" + json.dumps(metrics, indent=4, sort_keys=True) + CEND)
    metric = input (Cyan + "which one you want to query? (copy and paste here): " +  CEND)
    #print (Cyan + "Here are the available dimensions: \n" + json.dumps(metrics, indent=4, sort_keys=True))
    #dimension = input (Cyan + "which one you want to query? (copy and paste here) ")
    req_payload = {
    "t0": t0,    
    "t1": t1,    
    "dimension": "src_port",
    "metric": metric,
    #"filter": {"type": "eq", "field": "src_address", "value": "172.29.203.193"},  #optional
    "threshold": int(threshold),
    "scopeName": scope_name
    }


    resp = rc.post('/flowsearch/topn',
               json_body=json.dumps(req_payload))

    #print (json.dumps(cve_hosts, indent=4))
    if resp.status_code != 200:
        print(URED + "Failed to retrieve TopN")
        print(resp.status_code)
        print(resp.text)
    else:
        topN = resp.json()
        #print (json.dumps(topN, indent=4))
        topN_list = []   # store TopN data
        print (json.dumps(topN[0]['result'], indent=4))
        for top in topN[0]['result']:
            topN_dict = {}
            topN_dict["Source Service"] = top["src_port"]
            topN_dict[metric] = top[metric]
            topN_list.append(topN_dict)

        # specify csv file for exporting
        export_csvfile = 'topSrcPort.xlsx'

        # specify csv header fields
        csv_header = ["Source Service", metric]

        
        workbook = xlsxwriter.Workbook(export_csvfile)
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet(name='Top Source Service')
        cell_format = workbook.add_format()
        cell_format.set_bg_color('cyan')
        cell_format.set_bold()
        cell_format.set_font_color('black')
        worksheet.set_row(0, None)
        worksheet.write_row(0,0,csv_header,cell_format)
        i=1
        for row in topN_list:
            #print (row.values())
            worksheet.write_row(i,0,row.values())
            i+=1
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        workbook.close()
        print ('Writing csv file to %s with %d columns' % (export_csvfile, len(csv_header)))

def filterToString(invfilter):
    if 'filters' in invfilter.keys():
        query=[]
        for x in invfilter['filters']:
            if 'filters' in x.keys():
                query.append(filterToString(x))
            elif 'filter' in x.keys():
                query.append(x['type'] + filterToString(x['filter']))
            else:
                query.append(x['field'].replace('user_','*')+ ' '+ x['type'] + ' '+ str(x['value']))
        operator = ' '+invfilter['type']+' '
        return '('+operator.join(query)+')'
    else:
        return invfilter['field']+ ' '+ invfilter['type'] + ' '+ str(invfilter['value'])

def selectTetApps(apps):
    # Return App IDa for one or many Tetration Apps that we choose
    print (Cyan + "\nHere are all Application workspaces in your cluster: " + CEND)
    ShowApps(apps)
    choice = input('\nSelect which Tetration Apps (Number, Number) above you want to download polices: ')

    choice = choice.split(',')
    appIDs = []
    for app in choice:
        if '-' in app:
            for app in range(int(app.split('-')[0])-1,int(app.split('-')[1])):
                appIDs.append(resp.json()[int(app)-1]['id'])
        else:
            appIDs.append(apps[int(app)-1]['id'])
    return appIDs

def downloadPolicies(rc,appIDs):
    # Download Policies JSON files from Apps workspace
    apps = []
    for appID in appIDs:
        print('Downloading app details for '+appID + "into json file")
        apps.append(rc.get('/openapi/v1/applications/%s/details'%appID).json())
        #json_object = json.load(apps)
    for app in apps:
        with open('./'+app['name'].replace('/','-')+'.json', "w") as config_file:
            json.dump(apps, config_file, indent=4)
            print(app['name'].replace('/','-')+".json created")
    return apps

def convApps2xls(rc):
    AllApps = GetApps(rc)
    scopes = GetApplicationScopes(rc)
    apps = []
    appIDs = selectTetApps(AllApps)
    apps.append(downloadPolicies(rc, appIDs))
    print (BLINK + CRED +'Processing Application data ........ '+ CEND)
    #print (json.dumps(apps, indent=4))

    # Load in the IANA Protocols
    protocols = {}
    try: 
        with open('protocol-numbers-1.csv') as protocol_file:
            reader = csv.DictReader(protocol_file)
            for row in reader:
                protocols[row['Decimal']]=row
    except IOError:
        print('%% Could not load protocols file')
        return
    except ValueError:
        print('Could not load improperly formatted protocols file')
        return
    
    for app in apps[0]:
        workbook = xlsxwriter.Workbook(app['name'].replace('/','-')+'.xlsx')
        bold = workbook.add_format({'bold': True})

        if 'clusters' in app.keys():
            worksheet = workbook.add_worksheet(name='App Servers')
            cell_format = workbook.add_format()
            cell_format.set_bg_color('cyan')
            cell_format.set_bold()
            cell_format.set_font_color('black')
            worksheet.set_row(0, None)
            worksheet.write_row(0,0,['Hostname','IP','Cluster Membership'],cell_format)
            i=1
            clusters = app['clusters']
            for cluster in clusters:
                hosts = []
                for node in cluster['nodes']:
                    hosts.append(node['name'])
                    worksheet.write_row(i,0,[node['name'],node['ip'],cluster['name']])
                    i+=1
            worksheet.set_column(0, 0, 30)
            worksheet.set_column(1, 1, 15)
            worksheet.set_column(2, 2, 30)

        if 'inventory_filters' in app.keys():
            i=1
            worksheet = workbook.add_worksheet(name='External Groups')
            cell_format = workbook.add_format()
            cell_format.set_text_wrap()
            header_format = workbook.add_format()
            header_format.set_bg_color('cyan')
            header_format.set_bold()
            header_format.set_font_color('black')
            worksheet.set_row(0, None)
            worksheet.write_row(0,0,['Inventory Filter Name', 'IP Addresses', 'Filter Definition'],header_format)
            worksheet.set_column(0, 0, 30)
            worksheet.set_column(1, 1, 60, cell_format)
            worksheet.set_column(2, 2, 50, cell_format)

            filters = app['inventory_filters']
            for invfilter in filters:
                #print (json.dumps(invfilter, indent=4))
                ipSet = resolveFilter(rc, invfilter)
                #print (ipSet)
                worksheet.write_row(i,0,[invfilter['name'], str(ipSet), filterToString(invfilter['query'])])
                i+=1

        if 'default_policies' in app.keys():
            i=1
            worksheet = workbook.add_worksheet(name='Policies')
            header_format = workbook.add_format()
            header_format.set_bg_color('cyan')
            header_format.set_bold()
            header_format.set_font_color('black')
            worksheet.set_row(0, None)
            worksheet.write_row(0,0,['Consumer Group','Provider Group','Services'],header_format)
            worksheet.set_column(0, 0, 30)
            worksheet.set_column(1, 1, 30)
            worksheet.set_column(2, 2, 30)

            policies = app['default_policies']
            for policy in policies:
                pols = {}
                for rule in policy['l4_params']:
                    if 'port' in rule:
                        if rule['port'][0] == rule['port'][1]:
                            port = str(rule['port'][0])
                        else:
                            port = str(rule['port'][0]) + '-' + str(rule['port'][1])
                    else:
                        port = None

                    if port == None:
                        try:
                            pols[protocols[str(rule['proto'])]['Keyword']] = []
                        except:
                            pols['PROTO-'+str(rule['proto'])]=[]
                    elif protocols[str(rule['proto'])]['Keyword'] in pols.keys():
                        pols[protocols[str(rule['proto'])]['Keyword']].append(port)
                    else:
                        pols[protocols[str(rule['proto'])]['Keyword']] = [port]

                policy_list = []
                for key, val in pols.items():
                    #print(key,val)
                    if len(val)>0:
                        policy_list.append('{}={}'.format(key,', '.join(val)))
                    else:
                        policy_list.append(key)
                        
                worksheet.write_row(i,0,[policy["consumer_filter_name"],policy["provider_filter_name"],'; '.join(policy_list)])
                i+=1
        
        workbook.close()
        print (app['name'].replace('/','-')+'.xlsx created for policies conversion to CSV')


def resolveFilter(rc, filters):# return all IP and hosts for a specific filters
    ipSet = []
    body = json.dumps({'filter':filters['query']})
   
    resp = rc.post('/inventory/search',json_body=body)
    if resp:
        ips = resp.json()
        for i in ips['results']:
            ipSet.append(i['ip'])

    return ipSet

def main():
    print (BLINK + BOLD+ CGREEN + "Welcome to Tetration Report !!!" + CEND)
    print (BOLD+ Cyan+ UNDERLINE + ITALIC +"Build report for Tetration, Command: report and sub command: workloads or flows or apps. Use ? or help for more information"+ CEND)
    rc = CreateRestClient()
    command = input ("tetcli #  "+ CEND)

    # report 
    if command == "report h" or command =="report help" or command =="report ?": 
        print (BOLD+ CYELLOW + "Build report for Tetration, sub command: workloads or flows or apps  "+ CEND)
    if command == "report workloads" or command == "report wl" or command == "report workloads ?" or command == "report workloads h" or command == "report workloads help" or command == "report wl ?" or command == "report wl h" or command == "report wl help":
        print (BOLD+ CYELLOW + "Build report for Tetration workloads, sub command: all or detail or stats or software or vulnerabilities or processes "+ CEND)
        print (BOLD+ CYELLOW + "All - Report all installed workloads in your cluster in all scopes  "+ CEND)
        print (BOLD+ CYELLOW + "Detail - Detail Report about a specific workload  "+ CEND)
        print (BOLD+ CYELLOW + "Stats - Detail Workload communication report from time (t0) to time(t1)  "+ CEND)
        print (BOLD+ CYELLOW + "Software - Detail Installed Software Packages report for a specific workload  "+ CEND)
        print (BOLD+ CYELLOW + "Vulnerabilities - Detail Vulnerable Software Packages report for a specific workload or all workloads that match a CVE Score query. Sub: workload or all  "+ CEND)
        print (BOLD+ CYELLOW + "Processes - Detail Running processes report for a specific workload. Sub command: summary or all  "+ CEND)
    if command == "report flows" or command == "report flow" or command == "report flows ?" or command == "report flows h" or command == "report flows help" or command == "report flow ?" or command == "report flow h" or command == "report flow help":
        print (BOLD+ CYELLOW + "inv - Detail flow communication report about a subnet in a VRF from time (t0) to time(t1) "+ CEND)
        print (BOLD+ CYELLOW + "top - Top Talkers/Destination/Service report in excel for a scope from time (t0) to time(t1). Sub command: talkers, servers, sports, dports "+ CEND)
    if command == "report apps" or command == "report app" or command == "report apps ?" or command == "report apps h" or command == "report apps help" or command == "report app ?" or command == "report app h" or command == "report app help":
        print (BOLD+ CYELLOW + "Build report for Tetration Apps, sub command: policies or conversation "+ CEND)
        print (BOLD+ CYELLOW + "Policies - Report policies in xlsx format for a specific Application  "+ CEND)
        print (BOLD+ CYELLOW + "Conversation - Report conversation in xlsx format for a specific Application  "+ CEND)
    if command == "report workloads all" or command == "report wl all" or command == "report workloads a" or command == "report wl a": 
        sensors = GetSensors(rc)
        print (BOLD+ CYELLOW + "\nHere are all Software Sensors in your cluster: " + CEND)
        ShowAgents(sensors)
    if command == "report workloads detail" or command == "report wl detail" or command == "report workloads det" or command == "report wl det": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        agent = GetAgentProfile(rc,uuid)
        ShowAgentProfile(agent)
    if command == "report workloads stats" or command == "report wl stats" or command == "report workloads st" or command == "report wl st": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        from_year = input(CYELLOW + "From which year (yyyy) you want to query: "+CEND)
        from_month = input(CYELLOW + "Month (mm)? "+CEND)
        from_day = input(CYELLOW + "Day (dd)? "+CEND)
        to_year = input(CYELLOW + "To which year (yyyy) you want to query: "+CEND)
        to_month = input(CYELLOW + "Month (mm)? "+CEND)
        to_day = input(CYELLOW + "Day (dd)? "+CEND)
        td = input(CYELLOW + "What is the granularity (day, hour or minute)? "+CEND)
        t0 = round(datetime.datetime(int(from_year),int(from_month),(int(from_day)+1),0,0).timestamp())
        t1 = round(datetime.datetime(int(to_year),int(to_month),(int(to_day)+1),0,0).timestamp())
        stats = GetWorkloadStats(rc,uuid,t0,t1,td)
        print ("Here is the detail communication for your agent with UUID: " + uuid + " from " + from_day + "/"+ from_month + "/"+ from_year+ " to " + to_day + "/"+ to_month + "/"+ to_year)
        ShowWorkloadStats(stats)
    if command == "report workloads software" or command == "report wl software" or command == "report workloads sw" or command == "report wl sw": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        packages = GetSwPackages(rc,uuid)
        print ("Here are all the software packages installed in your agent with UUID: " + uuid)
        ShowSwPackages(packages)
    if command == "report workloads vulnerabilities" or command == "report wl vulnerabilities" or command == "report workloads vul" or command == "report wl vul": 
        print (BOLD+ CYELLOW + "Vulnerabilities - Detail Vulnerable Software Packages report for a specific workload or all workloads that match a CVE Score query. Sub: workload or all  "+ CEND)
        print (BOLD+ CYELLOW + "Please choose workloads or all as subcommand"+ CEND)
    if command == "report workloads vulnerabilities all" or command == "report wl vulnerabilities all" or command == "report workloads vul all" or command == "report wl vul all": 
        get_inventory_cve(rc)
    if command == "report workloads vulnerabilities workloads" or command == "report wl vulnerabilities workloads" or command == "report workloads vul wl" or command == "report wl vul wl": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        vuls = GetVul(rc,uuid)
        print ("Here are all vulnerable packages installed in your agent with UUID: " + uuid)
        ShowVul(vuls)
    if command == "report workloads processes" or command == "report wl processes" or command == "report workloads proc" or command == "report wl proc": 
        print (BOLD+ CYELLOW + "Processes - Detail Running processes report for a specific workload. Sub command: summary or all  "+ CEND)
        print (BOLD+ CYELLOW + "Please choose summary or all as subcommand"+ CEND)
    if command == "report workloads processes all" or command == "report wl processes all" or command == "report workloads proc all" or command == "report wl proc all": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        proc = GetProc(rc,uuid)
        print ("Here are all long running processes in your agent with UUID: " + uuid)
        ShowProc(proc)
    if command == "report workloads processes summary" or command == "report wl processes summary" or command == "report workloads proc sum" or command == "report wl proc sum": 
        sensors = GetSensors(rc)
        uuid = selectAgent(sensors)
        proc = GetProcTree(rc,uuid)
        handle = proc['process_summary'][0]['summary'][0]['handle']
        procDetail = GetProcTreeDetail(rc,uuid, handle)
        print ("Here are  process snapshot detail in your agent with UUID: " + uuid)
        #print (json.dumps(procDetail, indent=4))
        ShowProcTreeDetail(procDetail)
    if command == "report flow inventories" or command == "report flow inv" or command == "report flows inv" or command == "report flows inventories":
        get_inventory_flow(rc)
    if command == "report flow top" or command == "report flows top" or command == "report flow top ?" or command == "report flows top ?":
        print (BOLD+ CYELLOW + "top - Top Talkers/Destination/Service report in excel for a scope from time (t0) to time(t1). Sub command: talkers, servers, cservice, dservice "+ CEND)
        print (BOLD+ CYELLOW + "Please provide Sub command: talkers, servers, sports, dports "+ CEND)
    if command == "report flow top talkers" or command == "report flows top talkers" or command == "report flow top t" or command == "report flows top t":
        get_flow_topTalkers(rc)
    if command == "report flow top servers" or command == "report flows top servers" or command == "report flow top s" or command == "report flows top s":
        get_flow_topDest(rc)
    if command == "report flow top sports" or command == "report flows top sports" or command == "report flow top sp" or command == "report flows top sp":
        get_flow_topSrcService(rc)
    if command == "report flow top dports" or command == "report flows top dports" or command == "report flow top dp" or command == "report flows top dp":
        get_flow_topDestService(rc)
    if command == "report apps policies" or command == "report app policies" or command == "report apps pol" or command == "report app pol":
        convApps2xls(rc)
    if command == "report apps conversation" or command == "report app conversation" or command == "report apps conv" or command == "report app conv":
        AllApps = GetApps(rc)
        appIDs = selectTetApps(AllApps)
        downloadConvs(rc, appIDs)
        with open('all-conversations.json') as config_file:
            ShowConversationTet(json.load(config_file))


if __name__ == "__main__":
    main()